"""Benchmark: FlashInfer fused MoE kernels (tp8, ep8 config).

Tests four FlashInfer trtllm fused MoE variants:
  1. fp8_per_tensor_fuse_moe — trtllm_fp8_per_tensor_scale_moe
  2. mxfp8_mxfp8_fuse_moe   — trtllm_fp8_block_scale_moe with Fp8QuantizationType.MxFp8
  3. mxfp8_mxfp4_fuse_moe   — trtllm_fp4_block_scale_moe with mxfp8 activation + mxfp4 weight
  4. nvfp4_fuse_moe          — cute_dsl_fused_moe_nvfp4 with nvfp4 activation + nvfp4 weight

All use routing internally (trtllm APIs compute routing from logits).
Only expert computation latency is measured after warmup.

Usage:
    python tests/bench_flasinfer_fuse_moe.py
"""

import torch
import torch.cuda.nvtx as nvtx

from flashinfer import (
    mxfp8_quantize,
    fp4_quantize,
    nvfp4_quantize,
    shuffle_matrix_a,
    shuffle_matrix_sf_a,
)
from flashinfer.fused_moe import (
    cute_dsl_fused_moe_nvfp4,
    reorder_rows_for_gated_act_gemm,
    trtllm_fp4_block_scale_moe,
    trtllm_fp8_block_scale_moe,
    trtllm_fp8_per_tensor_scale_moe,
)
from flashinfer.tllm_enums import Fp8QuantizationType


# ---------------------------- config ----------------------------

NUM_EXPERT_TOTAL = 256
NUM_TOPK = 8

# Total token counts (m = num_seq) to sweep.
M_CASES = [256, 320, 384, 448, 512, 1024, 16384, 32768]
M_CASES = [32768 * 2]

# (label, hidden, intermediate_size, num_expert_local, rank_ep)
SHAPES = [
    # ("h6144_tp8", 6144, 256, 256, 0),
    # ("h6144_tp4", 6144, 512, 256, 0),
    ("hy4_ep8", 6144, 2048, 32, 0),
    ("hy4_tp8", 6144, 256, 256, 0),
    # ("h4096_tp8", 4096, 192, 192, 0),
    # ("h4096_tp4", 4096, 384, 192, 0),
    # ("hy3_ep8", 4096, 1536, 24, 0),
]

SF_VEC_MXFP = 32  # scale factor vector size for mxfp8/mxfp4
WARMUP = 3
ITER = 10


# ---------------------------- helpers ----------------------------


def _build_routing_logits(num_seq, device):
    """Build routing logits for trtllm APIs (routing is computed internally)."""
    routing_logits = torch.randn(num_seq, NUM_EXPERT_TOTAL, dtype=torch.float32, device=device)
    return routing_logits


# ---------------------------- builder: mxfp8 x mxfp8 ----------------------------


def _build_mxfp8_mxfp8(hidden, inter, num_expert_local, rank_ep, num_seq, device):
    """trtllm_fp8_block_scale_moe with Fp8QuantizationType.MxFp8.

    MxFp8 activation x MxFp8 weight, block size = 32.
    Weights and scales must be shuffled for the kernel (use_shuffled_weight=True).
    """
    if hidden % 128 != 0 or inter % 128 != 0:
        return None
    gate_up_n = inter * 2
    epilogue_tile_m = 128  # MxFp8 uses epilogue_tile_m=128

    # Generate bf16 source then quantize activation to mxfp8
    x_bf16 = torch.randn((num_seq, hidden), dtype=torch.bfloat16, device=device) / 100
    mxfp8_x, mxfp8_x_sf = mxfp8_quantize(x_bf16, False)  # linear layout
    hidden_states_scale = mxfp8_x_sf.reshape(num_seq, hidden // SF_VEC_MXFP).contiguous()

    # Weights: quantize to mxfp8 per expert
    w1_bf16 = (
        torch.randn((num_expert_local, gate_up_n, hidden), dtype=torch.bfloat16, device=device) / 10
    )
    w2_bf16 = (
        torch.randn((num_expert_local, hidden, inter), dtype=torch.bfloat16, device=device) / 10
    )

    # Quantize weights with mxfp8_quantize (per expert)
    gemm1_weights_list = []
    gemm1_scales_list = []
    gemm2_weights_list = []
    gemm2_scales_list = []

    for i in range(num_expert_local):
        w1_fp8, w1_sf = mxfp8_quantize(w1_bf16[i], False)
        w2_fp8, w2_sf = mxfp8_quantize(w2_bf16[i], False)

        # Reshape scales: mxfp8_quantize returns flat uint8, reshape to [N, K//32]
        w1_sf = w1_sf.reshape(gate_up_n, hidden // SF_VEC_MXFP)
        w2_sf = w2_sf.reshape(hidden, inter // SF_VEC_MXFP)

        # Reorder rows for gated activation (SwiGLU) on gemm1
        w1_fp8_reordered = reorder_rows_for_gated_act_gemm(w1_fp8)
        w1_sf_reordered = reorder_rows_for_gated_act_gemm(w1_sf)

        # Shuffle weight matrices
        w1_shuffled = shuffle_matrix_a(w1_fp8_reordered.view(torch.uint8), epilogue_tile_m)
        w2_shuffled = shuffle_matrix_a(w2_fp8.view(torch.uint8), epilogue_tile_m)

        # Shuffle scale matrices
        w1_sf_shuffled = shuffle_matrix_sf_a(w1_sf_reordered, epilogue_tile_m)
        w2_sf_shuffled = shuffle_matrix_sf_a(w2_sf, epilogue_tile_m)

        gemm1_weights_list.append(w1_shuffled)
        gemm2_weights_list.append(w2_shuffled)
        gemm1_scales_list.append(w1_sf_shuffled)
        gemm2_scales_list.append(w2_sf_shuffled)

    gemm1_weights = torch.stack(gemm1_weights_list).view(torch.float8_e4m3fn)
    gemm2_weights = torch.stack(gemm2_weights_list).view(torch.float8_e4m3fn)
    gemm1_weights_scale = torch.stack(gemm1_scales_list)  # [E, N*K/32] flat
    gemm2_weights_scale = torch.stack(gemm2_scales_list)  # [E, N*K/32] flat

    routing_logits = _build_routing_logits(num_seq, device)

    def run():
        return trtllm_fp8_block_scale_moe(
            routing_logits=routing_logits,
            routing_bias=None,
            hidden_states=mxfp8_x,
            hidden_states_scale=hidden_states_scale,
            gemm1_weights=gemm1_weights,
            gemm1_weights_scale=gemm1_weights_scale,
            gemm2_weights=gemm2_weights,
            gemm2_weights_scale=gemm2_weights_scale,
            num_experts=NUM_EXPERT_TOTAL,
            top_k=NUM_TOPK,
            n_group=None,
            topk_group=None,
            intermediate_size=inter,
            local_expert_offset=rank_ep * num_expert_local,
            local_num_experts=num_expert_local,
            routed_scaling_factor=None,
            use_shuffled_weight=True,
            fp8_quantization_type=Fp8QuantizationType.MxFp8,
        )

    # Trigger JIT/cubin load eagerly
    try:
        run()
        torch.cuda.synchronize()
    except Exception as e:
        print(f"  [SKIP] mxfp8_mxfp8: {e}")
        return None

    return run


# ---------------------------- builder: mxfp8 x mxfp4 ----------------------------


def _build_mxfp8_mxfp4(hidden, inter, num_expert_local, rank_ep, num_seq, device):
    """trtllm_fp4_block_scale_moe with mxfp8 activation + mxfp4 weight.

    Activation: mxfp8 (sf_vec=32, ue8m0 scales).
    Weight: mxfp4 packed (sf_vec=32, ue8m0 scales).
    """
    if hidden % 128 != 0 or inter % 128 != 0:
        return None
    gate_up_n = inter * 2

    # Generate bf16 source tensors then quantize
    x_bf16 = torch.randn((num_seq, hidden), dtype=torch.bfloat16, device=device) / 100
    w1_bf16 = (
        torch.randn((num_expert_local, gate_up_n, hidden), dtype=torch.bfloat16, device=device) / 10
    )
    w2_bf16 = (
        torch.randn((num_expert_local, hidden, inter), dtype=torch.bfloat16, device=device) / 10
    )

    # Quantize activation: mxfp8 (linear layout for trtllm)
    mxfp8_x, mxfp8_x_sf = mxfp8_quantize(x_bf16, False)
    mxfp8_x_sf = mxfp8_x_sf.view(torch.float8_e4m3fn).reshape(num_seq, -1)

    # Quantize weights: mxfp4 via fp4_quantize (sf_vec_size=32, sf_use_ue8m0=True)
    w1_fp4, w1_sf = fp4_quantize(
        w1_bf16, torch.tensor([1.0], device=device), sf_vec_size=32, sf_use_ue8m0=True
    )
    w1_sf = w1_sf.view(torch.float8_e4m3fn).reshape(num_expert_local, gate_up_n, -1)

    w2_fp4, w2_sf = fp4_quantize(
        w2_bf16, torch.tensor([1.0], device=device), sf_vec_size=32, sf_use_ue8m0=True
    )
    w2_sf = w2_sf.view(torch.float8_e4m3fn).reshape(num_expert_local, hidden, -1)

    # Output scale scalars
    output1_scale = torch.ones(num_expert_local, device=device)
    output1_gate_scale = torch.ones(num_expert_local, device=device)
    output2_scale = torch.ones(num_expert_local, device=device)

    routing_logits = _build_routing_logits(num_seq, device)

    def run():
        return trtllm_fp4_block_scale_moe(
            routing_logits,
            None,  # routing_bias
            mxfp8_x,
            mxfp8_x_sf,
            w1_fp4,
            w1_sf,
            None,  # gemm1_bias
            None,  # gemm1_alpha
            None,  # gemm1_beta
            None,  # gemm1_clamp_limit
            w2_fp4,
            w2_sf,
            None,  # gemm2_bias
            output1_scale,
            output1_gate_scale,
            output2_scale,
            NUM_EXPERT_TOTAL,
            NUM_TOPK,
            None,  # n_group
            None,  # topk_group
            inter,
            rank_ep * num_expert_local,  # local_expert_offset
            num_expert_local,
            None,  # routed_scaling_factor
        )

    # Trigger JIT/cubin load eagerly
    try:
        run()
        torch.cuda.synchronize()
    except Exception as e:
        print(f"  [SKIP] mxfp8_mxfp4: {e}")
        return None

    return run


# ---------------------------- builder: fp8 per-tensor ----------------------------


def _build_fp8_per_tensor(hidden, inter, num_expert_local, rank_ep, num_seq, device):
    """trtllm_fp8_per_tensor_scale_moe (fp8 per-tensor, includes routing)."""
    gate_up_n = inter * 2

    # FP8 per-tensor: quantize hidden_states and weights to fp8
    x_bf16 = torch.randn((num_seq, hidden), dtype=torch.bfloat16, device=device) / 100
    x_fp8 = x_bf16.to(torch.float8_e4m3fn)

    w1_bf16 = (
        torch.randn((num_expert_local, gate_up_n, hidden), dtype=torch.bfloat16, device=device) / 10
    )
    w2_bf16 = (
        torch.randn((num_expert_local, hidden, inter), dtype=torch.bfloat16, device=device) / 10
    )
    w1_fp8 = w1_bf16.to(torch.float8_e4m3fn)
    w2_fp8 = w2_bf16.to(torch.float8_e4m3fn)

    # Per-expert output scales
    output1_scale = torch.full((num_expert_local,), 0.25, dtype=torch.float32, device=device)
    output1_gate_scale = torch.full((num_expert_local,), 0.25, dtype=torch.float32, device=device)
    output2_scale = torch.full((num_expert_local,), 0.25, dtype=torch.float32, device=device)

    routing_logits = _build_routing_logits(num_seq, device)

    def run():
        return trtllm_fp8_per_tensor_scale_moe(
            routing_logits,
            None,  # routing_bias
            x_fp8,
            w1_fp8,
            output1_scale,
            output1_gate_scale,
            w2_fp8,
            output2_scale,
            NUM_EXPERT_TOTAL,
            NUM_TOPK,
            None,  # n_group
            None,  # topk_group
            inter,
            rank_ep * num_expert_local,  # local_expert_offset
            num_expert_local,
            None,  # routed_scaling_factor
            False,  # use_routing_scales_on_input
        )

    # Trigger JIT/cubin load eagerly
    try:
        run()
        torch.cuda.synchronize()
    except Exception as e:
        print(f"  [SKIP] fp8_per_tensor: {e}")
        return None

    return run


# ---------------------------- builder: nvfp4 x nvfp4 ----------------------------

SF_VEC_NVFP4 = 16
FLOAT8_E4M3_MAX = 448.0
FLOAT4_E2M1_MAX = 6.0


def _build_nvfp4_nvfp4(hidden, inter, num_expert_local, rank_ep, num_seq, device):
    """cute_dsl_fused_moe_nvfp4 with NVFP4 activation + NVFP4 weight.

    Uses pre-computed routing (cute_dsl API takes topk_ids/scales directly).
    """
    if hidden % 128 != 0 or inter % 128 != 0:
        return None
    gate_up_n = inter * 2
    quant_blocksize = SF_VEC_NVFP4

    def round_up(x_val, y):
        return (x_val + y - 1) // y * y

    # Generate source tensors
    x_bf16 = torch.randn((num_seq, hidden), dtype=torch.bfloat16, device=device) / 100
    w1_bf16 = (
        torch.randn((num_expert_local, gate_up_n, hidden), dtype=torch.bfloat16, device=device) / 10
    )
    w2_bf16 = (
        torch.randn((num_expert_local, hidden, inter), dtype=torch.bfloat16, device=device) / 10
    )

    # Quantize weights per expert to NVFP4
    w1_q = torch.empty((num_expert_local, gate_up_n, hidden // 2), device=device, dtype=torch.uint8)
    w2_q = torch.empty((num_expert_local, hidden, inter // 2), device=device, dtype=torch.uint8)
    w1_blockscale = torch.empty(
        (num_expert_local, round_up(gate_up_n, 128), round_up(hidden // quant_blocksize, 4)),
        device=device,
        dtype=torch.float8_e4m3fn,
    )
    w2_blockscale = torch.empty(
        (num_expert_local, round_up(hidden, 128), round_up(inter // quant_blocksize, 4)),
        device=device,
        dtype=torch.float8_e4m3fn,
    )
    w1_gs = torch.empty((num_expert_local,), device=device, dtype=torch.float32)
    w2_gs = torch.empty((num_expert_local,), device=device, dtype=torch.float32)

    for expert in range(num_expert_local):
        w1_src = w1_bf16[expert]
        w2_src = w2_bf16[expert].contiguous()
        w1_amax = torch.abs(w1_src).max().to(torch.float32)
        w2_amax = torch.abs(w2_src).max().to(torch.float32)
        w1_gs[expert] = FLOAT8_E4M3_MAX * FLOAT4_E2M1_MAX / w1_amax
        w2_gs[expert] = FLOAT8_E4M3_MAX * FLOAT4_E2M1_MAX / w2_amax
        w1_q[expert], w1_blockscale[expert] = fp4_quantize(w1_src, w1_gs[expert])
        w2_q[expert], w2_blockscale[expert] = fp4_quantize(w2_src, w2_gs[expert])

    # Activation global scale & quantize input to NVFP4
    a1_gs = torch.ones((), device=device, dtype=torch.float32)
    a2_gs = torch.ones((), device=device, dtype=torch.float32)
    hidden_states, input_sf = nvfp4_quantize(x_bf16, a1_gs)

    # Per-expert alpha = 1/(a_gs * w_gs)
    w1_alpha = 1.0 / (a1_gs * w1_gs)
    w2_alpha = 1.0 / (a2_gs * w2_gs)

    # fc2_input_scale: global scale for gemm2 input quantization
    fc2_input_scale = a2_gs

    # Pre-computed routing
    topk_ids = torch.randint(
        0, NUM_EXPERT_TOTAL, (num_seq, NUM_TOPK), device=device, dtype=torch.int32
    )
    topk_scale = torch.rand((num_seq, NUM_TOPK), device=device, dtype=torch.float32) * 0.5 + 0.5

    def run():
        return cute_dsl_fused_moe_nvfp4(
            x=hidden_states,
            x_sf=input_sf,
            token_selected_experts=topk_ids,
            token_final_scales=topk_scale,
            w1_weight=w1_q,
            w1_weight_sf=w1_blockscale,
            w1_alpha=w1_alpha,
            fc2_input_scale=fc2_input_scale,
            w2_weight=w2_q,
            w2_weight_sf=w2_blockscale,
            w2_alpha=w2_alpha,
            num_experts=NUM_EXPERT_TOTAL,
            top_k=NUM_TOPK,
            num_local_experts=num_expert_local,
            local_expert_offset=rank_ep * num_expert_local,
        )

    # Trigger JIT/cubin load eagerly
    try:
        run()
        torch.cuda.synchronize()
    except Exception as e:
        print(f"  [SKIP] nvfp4_nvfp4: {e}")
        return None

    return run


# ---------------------------- timing ----------------------------


def _bench(label: str, fn) -> float:
    """Returns mean per-call latency in microseconds (CUDA graph when possible)."""
    # Warmup (also triggers flashinfer autotuning)
    for _ in range(WARMUP):
        fn()
    torch.cuda.synchronize()

    # Try CUDA graph capture; fall back to eager if capture fails.
    graph = None
    try:
        graph = torch.cuda.CUDAGraph()
        with torch.cuda.graph(graph):
            fn()
        torch.cuda.synchronize()
    except Exception:
        graph = None
        torch.cuda.synchronize()

    # Timed loop
    nvtx.range_push(label)
    start = torch.cuda.Event(enable_timing=True)
    end = torch.cuda.Event(enable_timing=True)
    start.record()
    for _ in range(ITER):
        if graph is not None:
            graph.replay()
        else:
            fn()
    end.record()
    torch.cuda.synchronize()
    nvtx.range_pop()
    elapsed_ms = start.elapsed_time(end)

    # Release graph to avoid interfering with subsequent captures from other libraries.
    del graph
    torch.cuda.synchronize()

    return elapsed_ms * 1e3 / ITER  # us


# ---------------------------- main ----------------------------


def main():
    device = torch.device("cuda:0")
    torch.manual_seed(2026)

    print(
        f"NUM_EXPERT_TOTAL={NUM_EXPERT_TOTAL}  NUM_TOPK={NUM_TOPK}  "
        f"WARMUP={WARMUP}  ITER={ITER}"
    )

    # results[shape_label][num_seq] = (fp8_per_tensor_us, mxfp8_mxfp8_us, mxfp8_mxfp4_us)
    results = {label: {} for label, _, _, _, _ in SHAPES}

    for shape_label, hidden, inter, num_expert_local, rank_ep in SHAPES:
        print(
            f"\n=== shape={shape_label}  hidden={hidden}  intermediate={inter}  "
            f"num_expert_local={num_expert_local}/{NUM_EXPERT_TOTAL}  "
            f"topk={NUM_TOPK}  rank_ep={rank_ep} ===",
            flush=True,
        )
        for num_seq in M_CASES:
            num_seq_per_group = num_seq * NUM_TOPK // NUM_EXPERT_TOTAL

            # --- 1. fp8 per-tensor (trtllm_fp8_per_tensor_scale_moe) ---
            run_fp8_pt = _build_fp8_per_tensor(
                hidden, inter, num_expert_local, rank_ep, num_seq, device
            )
            if run_fp8_pt is None:
                t_fp8_pt = None
            else:
                t_fp8_pt = _bench(f"fp8_per_tensor/{shape_label}/n{num_seq}", run_fp8_pt)

            # --- 2. mxfp8 x mxfp8 (trtllm_fp8_block_scale_moe with MxFp8) ---
            run_mxfp8_mxfp8 = _build_mxfp8_mxfp8(
                hidden, inter, num_expert_local, rank_ep, num_seq, device
            )
            if run_mxfp8_mxfp8 is None:
                t_mxfp8_mxfp8 = None
            else:
                t_mxfp8_mxfp8 = _bench(f"mxfp8_mxfp8/{shape_label}/n{num_seq}", run_mxfp8_mxfp8)

            # --- 3. mxfp8 x mxfp4 (trtllm_fp4_block_scale_moe) ---
            run_mxfp8_mxfp4 = _build_mxfp8_mxfp4(
                hidden, inter, num_expert_local, rank_ep, num_seq, device
            )
            if run_mxfp8_mxfp4 is None:
                t_mxfp8_mxfp4 = None
            else:
                t_mxfp8_mxfp4 = _bench(f"mxfp8_mxfp4/{shape_label}/n{num_seq}", run_mxfp8_mxfp4)

            # --- 4. nvfp4 x nvfp4 (cute_dsl_fused_moe_nvfp4) ---
            run_nvfp4 = _build_nvfp4_nvfp4(
                hidden, inter, num_expert_local, rank_ep, num_seq, device
            )
            if run_nvfp4 is None:
                t_nvfp4 = None
            else:
                t_nvfp4 = _bench(f"nvfp4_nvfp4/{shape_label}/n{num_seq}", run_nvfp4)

            results[shape_label][num_seq] = (t_fp8_pt, t_mxfp8_mxfp8, t_mxfp8_mxfp4, t_nvfp4)

            def _f(t):
                return f"{t:7.2f}" if t is not None else "    N/A"

            print(
                f"  m={num_seq:>6d}  seq/grp={num_seq_per_group:>4d}  "
                f"fp8_per_tensor={_f(t_fp8_pt)}  mxfp8_mxfp8={_f(t_mxfp8_mxfp8)}  "
                f"mxfp8_mxfp4={_f(t_mxfp8_mxfp4)}  nvfp4={_f(t_nvfp4)}",
                flush=True,
            )

    # Export to Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    wb.remove(wb.active)

    for shape_label, hidden, inter, num_expert_local, rank_ep in SHAPES:
        ws = wb.create_sheet(title=shape_label)
        gate_up_n = inter * 2

        # Header
        headers = [
            "m",
            "seq/grp",
            "fp8pt(us)",
            "TFLOPS",
            "TB/s",
            "mxfp8x8(us)",
            "TFLOPS",
            "TB/s",
            "mxfp4(us)",
            "TFLOPS",
            "TB/s",
            "nvfp4(us)",
            "TFLOPS",
            "TB/s",
            "mxfp8x8/fp8pt",
            "mxfp4/fp8pt",
            "nvfp4/fp8pt",
        ]
        header_font = Font(bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, num_seq in enumerate(M_CASES, 2):
            num_seq_per_group = num_seq * NUM_TOPK // NUM_EXPERT_TOTAL
            t_fp8_pt, t_mxfp8_mxfp8, t_mxfp8_mxfp4, t_nvfp4 = results[shape_label][num_seq]

            m_eff = num_seq_per_group * num_expert_local
            flops = 2.0 * m_eff * gate_up_n * hidden + 2.0 * m_eff * hidden * inter
            act_bytes = m_eff * hidden * 2 + m_eff * hidden * 2

            w_bytes_fp8 = num_expert_local * (gate_up_n * hidden + hidden * inter) * 1
            bytes_fp8 = w_bytes_fp8 + act_bytes

            w_bytes_mxfp8x8 = num_expert_local * (gate_up_n * hidden + hidden * inter) * 1
            sf_bytes_mxfp8x8 = (
                num_expert_local * (gate_up_n * hidden // 32 + hidden * inter // 32) * 1
            )
            bytes_mxfp8x8 = w_bytes_mxfp8x8 + sf_bytes_mxfp8x8 + act_bytes

            w_bytes_mxfp4 = num_expert_local * (gate_up_n * hidden + hidden * inter) // 2
            sf_bytes_mxfp4 = (
                num_expert_local * (gate_up_n * hidden // 32 + hidden * inter // 32) * 1
            )
            bytes_mxfp4 = w_bytes_mxfp4 + sf_bytes_mxfp4 + act_bytes

            w_bytes_nvfp4 = num_expert_local * (gate_up_n * hidden + hidden * inter) // 2
            sf_bytes_nvfp4 = (
                num_expert_local * (gate_up_n * hidden // 16 + hidden * inter // 16) * 1
            )
            bytes_nvfp4 = w_bytes_nvfp4 + sf_bytes_nvfp4 + act_bytes

            def _tflops_val(t_us):
                if t_us is None or t_us <= 0:
                    return None
                return round(flops / (t_us * 1e-6) / 1e12, 1)

            def _bw_val(t_us, total_bytes):
                if t_us is None or t_us <= 0:
                    return None
                return round(total_bytes / (t_us * 1e-6) / 1e12, 2)

            def _speedup(base, t):
                if base is not None and t is not None and t > 0:
                    return round(base / t, 2)
                return None

            row = [
                num_seq,
                num_seq_per_group,
                round(t_fp8_pt, 2) if t_fp8_pt else None,
                _tflops_val(t_fp8_pt),
                _bw_val(t_fp8_pt, bytes_fp8),
                round(t_mxfp8_mxfp8, 2) if t_mxfp8_mxfp8 else None,
                _tflops_val(t_mxfp8_mxfp8),
                _bw_val(t_mxfp8_mxfp8, bytes_mxfp8x8),
                round(t_mxfp8_mxfp4, 2) if t_mxfp8_mxfp4 else None,
                _tflops_val(t_mxfp8_mxfp4),
                _bw_val(t_mxfp8_mxfp4, bytes_mxfp4),
                round(t_nvfp4, 2) if t_nvfp4 else None,
                _tflops_val(t_nvfp4),
                _bw_val(t_nvfp4, bytes_nvfp4),
                _speedup(t_fp8_pt, t_mxfp8_mxfp8),
                _speedup(t_fp8_pt, t_mxfp8_mxfp4),
                _speedup(t_fp8_pt, t_nvfp4),
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col, value=val)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    excel_path = "tests/bench_flashinfer_fuse_moe_results.xlsx"
    wb.save(excel_path)
    print(f"\nResults saved to {excel_path}")


if __name__ == "__main__":
    main()
